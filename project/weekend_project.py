import logging
import glob, os

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

# Making a dictionary for easy conversion from month name to month number
months = {
    'january': '01',
    'february': '02',
    'march': '03',
    'april': '04',
    'may': '05',
    'june': '06',
    'july': '07',
    'august': '08',
    'september': '09',
    'october': '10',
    'november': '11',
    'december': '12'
}

# Making program logger
logs = logging.getLogger('weekend_project')
logs.setLevel(logging.DEBUG)

# Configuring log file
file = logging.FileHandler('expedia_report.log')
file.setLevel(logging.NOTSET)

# Configuring console logger
console = logging.StreamHandler()
console.setLevel(logging.NOTSET)

# Setting up formatter for file and console logger
formatter = logging.Formatter('%(asctime)s [%(levelname)s]: %(message)s', '%m-%d-%Y %H:%M:%S')
file.setFormatter(formatter)
console.setFormatter(formatter)

# Linking file and console logger to program logger
logs.addHandler(file)
logs.addHandler(console)

# Start of log
logs.info('Program Started...')

# Setting active directory to look in and displaying Excel files in it
os.chdir('.')
logs.info('Displaying Excel files within \'project/\'')
logs.info(glob.glob('*.xlsx'))

# Taking Excel file name as input
logs.info('Querying for file name...')
name = ''
while True:
    print('Enter the file that you are trying to open without the extension.')
    name = str(input())

    logs.info('Checking if ' + name + '.xlsx exists...')
    if os.path.exists(name + '.xlsx'):
        logs.info(name + '.xlsx found. Opening file...')
        break
    else:
        logs.info(name + '.xlsx not found... Trying again')
        print(name + '.xlsx does not exist. Please try again.')

# Starting OpenPyXL for Excel file reading
wb = load_workbook(name + '.xlsx')

# Setting desired sheet to 'Summary Rolling MoM'
ws = wb['Summary Rolling MoM']

# Splitting input name to get the year and month
parts = name.split('_')
year = parts[-1]
month = parts[-2]

# Find which row the necessary data is on
logs.info('Finding row number for summary data...')
row_num = 0
for row in range(2, 14):
    if str(ws['A' + str(row)].value)[0:7] == year + '-' + months[month]:
        row_num = row
        logs.info('Row found. Row #' + str(row))
        break

# Logging start of data collection
logs.info('Loading Rolling MoM data for ' + month[0:1].upper() + month[1::].lower() + '...')

# Gathering data from 'Summary Rolling MoM' and logging it.
summary = {
    'Calls Offered': '{:,}'.format(ws['B' + str(row_num)].value),
    'Abandon after 30s': '{:.2%}'.format(ws['C' + str(row_num)].value),
    'FCR': '{:.2%}'.format(ws['D' + str(row_num)].value),
    'DSAT': '{:.2%}'.format(ws['E' + str(row_num)].value),
    'CSAT': '{:.2%}'.format(ws['F' + str(row_num)].value)
}
summaryList = [(k, v) for k, v in summary.items()]
logs.info('Summary: ' + str(summaryList))

# Displaying the data in a user friendly format
for data in summaryList:
    print(data[0] + ': ' + data[1])
print()

# Setting desired sheet to 'VOC Rolling MoM'
ws = wb['VOC Rolling MoM']

# Find which column the necessary data is on
logs.info('Finding column letter for VOC data...')
col_letter = 'A'
for col in range(2, 20):
    char = get_column_letter(col)
    if (str(ws[char + str(1)].value)[0:7] == year + '-' + months[month]) or (str(ws[char + str(1)].value).lower() == month):
        col_letter = char
        logs.info('Column found. Column #' + str(col) + '(' + char + ')')
        break

# Gathering data from 'VOC Rolling MoM' and logging it.
voc = {
    'Base Size': '{:,}'.format(int(ws[col_letter + '3'].value)),
    'Promoters': 'Good' if int(ws[col_letter + '4'].value) > 200 else 'Bad',
    'Passives': 'Good' if int(ws[col_letter + '6'].value) > 200 else 'Bad',
    'Dectractors': 'Good' if int(ws[col_letter + '8'].value) > 200 else 'Bad',
    'Overall NPS %': '{:.2%}'.format(ws[col_letter + '13'].value),
    'Sat with Agent %': '{:.2%}'.format(ws[col_letter + '16'].value),
    'DSat with Agent %': '{:.2%}'.format(ws[col_letter + '19'].value)
}
vocList = [(k, v) for k, v in voc.items()]
logs.info('VOC: ' + str(vocList))

# Displaying the data in a user friendly format
for data in vocList:
    print(data[0] + ': ' + data[1])

# Closing of log
logs.info('Finished gathering data for ' + month[0:1].upper() + month[1::].lower() + '...')
logs.info('Program Stopping...')

# Blank space for next execution
logs.info('###########################################################################################################')
